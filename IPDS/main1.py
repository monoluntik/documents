from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from petrovich.main import Petrovich
from petrovich.enums import Case, Gender
from num2words import num2words
from docxtpl import DocxTemplate
from random import randint
import os

CURRENCY_CODE = {'usd':'2',
                 'rub':'0',
                 'euro':'3'}
def get_sum_str(summ, currency='usd'):
    data = requests.get(f"https://summa-propisyu.ru/?summ={summ}&vat=20&val={CURRENCY_CODE[currency]}&sep=1")
    page = BeautifulSoup(data.text, "lxml")
    sum_str = page.find('textarea', {'id':'result1'}).text
    return sum_str

liiist = [ 'dataIPDS/Пояснительное  пиьсмо.docx', 'dataIPDS/ПРИЛОЖЕНИЕ 12 Элит Логистик.docx', 'dataIPDS/ПРИЛОЖЕНИЕ 1 2000000.docx', 'dataIPDS/КОНТРАКТ_IP_Sharshenova_Meerim_Zholdoshevna_Ozar_SMOC_14032023.docx', 'dataIPDS/Дополнительное соглашение.docx']
def generate():
    wb = load_workbook(filename = 'dataIPDS/Книга1.xlsx')
    ws = wb.get_sheet_by_name('Лист1')
    products = []
    for x in range(13, 16):
        product = {
            'product_name_ru':ws[f'B{x}'].value,
            'product_amount':int(ws[f'E{x}'].value),
            'product_price_for_one':round(float(f"{ws[f'D{x}'].value}".replace(",", ".")), 2),
            'product_price':round(float(f"{ws[f'F{x}'].value}".replace(",", ".")), 2),
            'product_name_en':ws[f'C{x}'].value,
            'com':round(float(f"{ws[f'F{x}'].value}".replace(",", "."))*0.005, 2),
        }
        products.append(product)
    ttns = []
    for x in range(13, 6):
        ttn = {    'contract_ip_ozar_number':f"{ws[f'K{x}'].value}".replace(',', '.'),
                   'our_date':f"{ws[f'C2'].value}".replace(',', '.'),
                   'our_director':f"{ws[f'L{x}'].value}".replace(',', '.'),
                   'contract_ip_ozar_date':f"{ws[f'J{x}'].value}".replace(',', '.'),
                   'our_name':f"{ws[f'H{x}'].value}".replace(',', '.'),
                   'our_contract':f"{ws[f'B2'].value}".replace(',', '.'),
                   }
        doc = DocxTemplate('dataIPDS/Транспортная_заявка_Элит_Логистик_HUBEI_HONGYANG_APPAREL_CO_,_LTD.docx')
        print(doc.get_undeclared_template_variables())
        inner = f'IPDS/{str(datetime.now().timestamp()).split(".")[0]}{randint(500, 8888888888888888888888888888)}.docx'
        doc.render(ttn)
        doc.save(inner)

        
    my_data = {
        'contract_ip_ozar_number' : f"{ws['C2'].value}".replace(',', '.'),
        'contract_ip_ozar_date' : f"{ws['B2'].value}".replace(',', '.'), 
        'appendix_ip_ozar_number' : f"{ws['C3'].value}".replace(',', '.'),
        'appendix_ip_ozar_date' : f"{ws['B3'].value}".replace(',', '.'),
        'ip_name_en' : f"{ws['E8'].value}".replace(',', '.'),
        'ip_name_ru' : f"{ws['B8'].value}".replace(',', '.'),
        'ip_address_en' : f"{ws['F8'].value}".replace(',', '.'),
        'ip_address_ru' : f"{ws['C8'].value}".replace(',', '.'),
        'ip_tin' : f"{ws['G8'].value}".replace(',', '.'),
        'ip_patent' : f"{ws['D8'].value}".replace(',', '.'),
        'dop_number' : f"{ws['C4'].value}".replace(',', '.'),
        'dop_date' : f"{ws['B4'].value}".replace(',', '.'),
        'contract_ozar_elit_number' : f"{ws['C5'].value}".replace(',', '.'),
        'contract_ozar_elit_date' : f"{ws['B5'].value}".replace(',', '.'),
        'appendix_ozar_elit_number' : f"{ws['C6'].value}".replace(',', '.'), 
        'appendix_ozar_elit_date' : f"{ws['B6'].value}".replace(',', '.'),
        'products' : products,
        'total' : sum([i.get('product_price', i['com']) for i in products]),
        'total_str_ru' : get_sum_str(sum([i.get('product_price', i['com'])]) for i in products)
    }
    for i in liiist:
        doc = DocxTemplate(i)
        print(doc.get_undeclared_template_variables())
        inner = f'IPDS/{str(datetime.now().timestamp()).split(".")[0]}{randint(500, 8888888888888888888888888888)}.docx'
        doc.render(my_data)
        doc.save(inner)

# print(os.listdir('dataIPDS/'))
generate()