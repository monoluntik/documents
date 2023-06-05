from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from petrovich.main import Petrovich
from petrovich.enums import Case, Gender
from num2words import num2words
from docxtpl import DocxTemplate
from random import randint

def get_sum_str(summ, currency):
    data = requests.get(f"https://summa-propisyu.ru/?summ={summ}&vat=20&val={CURRENCY_CODE[currency]}&sep=1")
    page = BeautifulSoup(data.text, "lxml")
    sum_str = page.find('textarea', {'id':'result1'}).text
    return sum_str


def get_datv_name(name):
    try:
        p = Petrovich()
        name = name.split(' ')
        cased_lname = p.lastname(name[0], Case.DATIVE, gender=Gender.MALE)
        cased_fname = p.firstname(name[1], Case.DATIVE, gender=Gender.MALE)
        return ' '.join([cased_lname, cased_fname])
    except: return name


def get_short_name(name):
    try:
        name = name.split(' ')
        short_name = name[0] + ' ' + name[1][0].upper() + '.' + name[2][0].upper() + '.'
        return short_name
    except: return name



# количество, цена за еденицу, наименование, еденица измерения, вид упаковки, итого
Products = [['1955', '87.45833333', 'ПвВГнг(A)-LS 3х2,5-0,66', 'м', 'Гофроящик', '170981.0417'],
['69', '9324.116667', 'Дифференц. автоматический выключатель 16А/30 мА, 10кА,  DS201, 2Р, С16, 30тА, ', 'шт', 'Гофроящик', '643364.05'],
['39', '9324.116667', 'Дифференц. автоматический выключатель 16А/30 мА, 10кА,  DS201, 2Р, С16, 30тА, ', 'шт', 'Гофроящик', '363640.55'],
['54', '9324.116667', 'Дифференц. автоматический выключатель 16А/30 мА, 10кА,  DS201, 2Р, С16, 30тА, ', 'шт', 'Гофроящик', '503502.3'],
['30', '9324.116667', 'Дифференц. автоматический выключатель 16А/30 мА, 10кА,  DS201, 2Р, С16, 30тА, ', 'шт', 'Гофроящик', '279723.5'],
['30', '9324.116667', 'Дифференц. автоматический выключатель 16А/30 мА, 10кА,  DS201, 2Р, С16, 30тА, ', 'шт', 'Гофроящик', '279723.5'],
['30', '9324.116667', 'Дифференц. автоматический выключатель 16А/30 мА, 10кА,  DS201, 2Р, С16, 30тА, ', 'шт', 'Гофроящик', '279723.5'],
['470', '508.0333333', 'ПвВГнг(A)-FRLS 5х16', 'м', 'Гофроящик', '238775.6667'],
['550', '583.6333333', 'ПвВГнг(A)-FRLS 5х10', 'м', 'Гофроящик', '320998.3333'],
['1645', '60.38333333', 'ПвВГнг(A)-LS 3х1,5', 'м', 'Гофроящик', '99330.58333'],
['96', '5538.208333', 'Короб силовой КЭТ 1890х300х150', 'шт', 'Гофроящик', '531668'],
['378', '734.7', 'Лоток перфорированный 400х50, L3000, толщ. 1,0 мм, горячеоцинкованный, 3526010HDZ', 'шт', 'Гофроящик', '277716.6'],
['1134', '323.8666667', 'Профиль L-0,5м, ВРL4105', 'шт', 'Гофроящик', '367264.8'],
['470', '508.0333333', 'ПвВГнг(A)-FRLS 5х16', 'м', 'Гофроящик', '238775.6667'],]

Document = {
    'number':'59',# вбивается вручную
    'date':'22.05.2023', # вбивается вручную
    'contract':'ИМТ-19.03.2023', # вбивается вручную
    'contract_date':'18.04.2023', # вбивается вручную
    'appendix':'1', # вбивается вручную
    'appendix_date':'18.04.2023', # вбивается вручную
}

Shipper = {
    'name':'Общество с ограниченной ответственностью «Мастер Транс Компани»',
    'short_name':'ОсОО «Мастер Транс Компани»',
    'okpo':'31564989', 
    'director':'Сагынова Нурзат Жанышовна', 
    'street':'',
    'city':'',
    'country':'',
    'tin': '00709202210014', 
}


powerOfAttorney = {
    'driver':'Нурматов Акылбек Таалайбекович', # выбор водителя рандом из базы
    'car':'Lexus', # машина привязання к водителю из базы
    'car_number':'K1524L', # номер машины привязанные к водителю из базы
    'prava':'PL154213', # права водителя привязанные к водителю из базы
    'passport':'ID5161418', # паспорт водителя из базы
    'passport_date':'31.06.2019' # дата выдачи паспорта водителя из базы
}
powerOfAttorney = {
    'driver':'Жакыпов Нурлан Азизович', # выбор водителя рандом из базы
    'car':'SHACMAN L3000', # машина привязання к водителю из базы
    'car_number':'CE9144', # номер машины привязанные к водителю из базы
    'prava':'  ', # права водителя привязанные к водителю из базы
    'passport':'AC4728463', # паспорт водителя из базы
    'passport_date':'SRS from 12.05.2019' # дата выдачи паспорта водителя из базы
}
Consignee = {
    'name':'Общество с ограниченной ответственностью «Интеграция»',
    'short_name':'ООО "ИНТЕГРАЦИЯ"',
    'okpo':'443995223',
    'director':'Голикова Анастасия Евгеньевна',
    'street':'',
    'city':'',
    'country':'',
    'tin':'000000000000012',
}

Payer = {
    'name':'Общество с ограниченной ответственностью «Интеграция»',
    'short_name':'ООО "ИНТЕГРАЦИЯ"',
    'okpo':'443995223',
    'director':'Голикова Анастасия Евгеньевна',
    'street':'',
    'city':'',
    'country':'',
    'tin':'000000000000012', 
}

Seller = {
    'name':'Общество с ограниченной ответственностью «Мастер Транс Компани»',
    'short_name':'ОсОО «Мастер Транс Компани»',
    'okpo':'31564989', 
    'director':'Сагынова Нурзат Жанышовна', 
    'street':'',
    'city':'',
    'country':'',
    'tin': '00709202210014', 
}

CUR = {'RUB':['Российский рубль', '643', 'Рублей'],
 'USD':['Доллар США', '840', 'Долларов США'],
 'KGS':['Сом', '417', 'Сом'],
 'CYN':['Юань', '156', 'Юаней'], 
 'EUR':['Евро', '978', 'Евро'],}


UNITS = {'м метр':'006',
 'кг килограмм':'166',
 'л литр':'112',
 'шт штука':'796',
 'PC unit':'796'}

CURRENCY_CODE = {'usd':'2',
                 'rub':'0',
                 'euro':'3'}


class Documents:

    def __init__(self, currency='rub', document = Document, shipper=Shipper, seller=Seller, powerOfAttorney=powerOfAttorney, products=Products, consignee=Consignee, payer=Payer, shablons_dir='shablons/', distance_dir='hello/'):
        self.shipper_addres = f'{shipper["street"]} {shipper["city"]} {shipper["country"]}'
        self.seller_addres = f'{seller["street"]} {seller["city"]} {seller["country"]}'
        self.consignee_addres = f'{consignee["street"]} {consignee["city"]} {consignee["country"]}'
        self.payer_addres = f'{payer["street"]} {payer["city"]} {payer["country"]}'
        self.shablons_dir = shablons_dir
        self.distance_dir = distance_dir
        self.currency = currency.lower()
        self.document = document
        self.shipper = shipper
        self.powerOfAttorney = powerOfAttorney
        self.products = products
        self.consignee = consignee
        self.payer = payer
        self.documents_number = document['number']
        self.documents_date = document['date']
        self.powerOfAttorney_number = f"{''.join([i[0] for i in powerOfAttorney['driver'].split(' ')])}{self.documents_date.replace('.', '')}"
        self.seller = seller
        self.shipper_info = f'{self.shipper["name"]}, {self.shipper_addres}, {self.shipper["tin"]}'
        self.shipper_director_short_name = get_short_name(self.shipper['director'])
        self.consignee_info = f'{self.consignee["name"]}, {self.consignee_addres}, {self.consignee["tin"]}'
        self.payer_info = f'{self.payer["name"]}, {self.payer_addres}, {self.payer["tin"]}'
        self.total_sum = '{0:,.2f}'.format(float(sum([float(i[-1]) for i in self.products])), 2).replace(',', ' ').replace('.', ',')
        self.total_products = f'{len(self.products)} ({num2words(len(self.products), lang="ru")})'
        self.total_products_count = '{0:,}'.format(sum([int(i[0]) for i in self.products])).replace(',', ' ')
        self.driver_short_name = get_short_name(powerOfAttorney["driver"])
        self.consignee_director_short_name = get_short_name(self.consignee['director'])
        self.payer_director_short_name = get_short_name(self.payer['director'])
        self.seller_director_short_name = get_short_name(self.seller['director'])
        self.seller_info = f'{self.seller["name"]}, {self.seller_addres}, {self.seller["tin"]}'

    def generate_ttn(self):


        wb = load_workbook(filename = f'{self.shablons_dir}ТТН.xlsx')
        ws = wb.get_sheet_by_name('TDSheet')
        
        
        ws['CU4'] = self.documents_number # Номер ТТН
        ws['CT217'] = self.documents_number # Номер ТТН вторая страница
        ws['CU5'] = self.documents_date # Дата ТТН


        ws['P6'] = self.shipper_info # Грузоотрпавитель полное наименование организации, адрес, номер телефона, ИНН
        ws['F218'] = self.shipper_info
        ws['CU6'] = self.shipper["okpo"] # ОКПО Грузоотрпавитель
        
        ws['AM208'] = self.shipper_director_short_name
        ws['AM212'] = self.shipper_director_short_name
        ws['Y245'] = self.shipper_director_short_name
        ws['AM210'] = self.shipper_director_short_name

        ws['P8'] =  self.consignee_info # Грузополучатель полное наименование организации, адрес, номер телефона, ИНН 
        ws['CU7'] = self.consignee['okpo'] # ОКПО Грузополучатель
        
        ws['P10'] =  self.payer_info # Плательщик полное наименование организации, адрес, номер телефона, ИНН
        ws['CU9'] = self.payer['okpo'] # ОКПО Плательщик
        ws['J222'] = ws['P10'].value
        
        
        ws['AV228'] = self.consignee_addres
        
        ws['CH192'] = self.total_sum # Сумма итого
        ws['CH193'] = self.total_sum # Сумма итого
        ws['CV201'] = self.total_sum # Сумма итого
        ws['B206'] = get_sum_str(self.total_sum, self.currency) # Сумма итого прописью
        

        ws['H197'] = self.total_products #'количество наименований'
        ws['H195'] = self.total_products #'количество наименований'
        ws['J204'] = self.total_products #'количество наименований'


        ws['H200'] = f'{self.total_products_count} ({num2words(sum([int(i[0]) for i in self.products]), lang="ru")})'
        ws['AD192'] = self.total_products_count
        ws['AD193'] = self.total_products_count
        ws['BW192'] = self.total_products_count
        ws['BW193'] = self.total_products_count


        ws['BD203'] = f'{get_datv_name(powerOfAttorney["driver"])} от {self.shipper["short_name"]}' 
        ws['CL208'] = self.driver_short_name
        ws['F224'] = self.driver_short_name

        ws['CC202'] = self.documents_date
        ws['BK202'] = self.powerOfAttorney_number
        ws['CL210'] = self.consignee_director_short_name

        for index, product in enumerate(self.products):
            index += 15
            ws[f'AD{index}'] = '{0:,}'.format(int(product[0])).replace(',', ' ')
            ws[f'AK{index}'] = '{0:,.2f}'.format(float(product[1]), 2).replace(',', ' ').replace('.', ',')
            ws[f'AR{index}'] = product[2]
            ws[f'BG{index}'] = product[3]
            ws[f'BM{index}'] = product[4]
            ws[f'BW{index}'] = '{0:,}'.format(int(product[0])).replace(',', ' ')
            ws[f'CH{index}'] = '{0:,.2f}'.format(float(product[5]), 2).replace(',', ' ').replace('.', ',')
            
        ws['F220'] = powerOfAttorney['car']
        ws['BH220'] = powerOfAttorney['car_number']
        ws['AV224'] = powerOfAttorney['prava']
        ws['F228'] = self.shipper_addres

        wb.save(self.distance_dir+str(datetime.now().timestamp()).split('.')[0]+'ТТН.xlsx')


    def generate_upd(self):

        wb = load_workbook(filename = f'{self.shablons_dir}УПД.xlsx')
        ws = wb.get_sheet_by_name('TDSheet')    

        ws['P1'] = self.documents_number
        ws['Y1'] = self.documents_date

        ws['R4'] = self.seller['name']
        ws['R5'] = self.seller_addres
        ws['R6'] = self.seller['tin']

        ws['R7'] = self.shipper_info
        ws['R8'] = self.consignee_info


        ws['BE4'] = self.payer['name']
        ws['BE5'] = self.payer_addres
        ws['BE6'] = self.payer['tin']

        ws['AC194'] = self.payer_director_short_name

        ws['BN194'] = self.payer_director_short_name

        ws['Z205'] = self.shipper_director_short_name

        ws['O207'] = self.documents_date

        ws['BM205'] = self.consignee_director_short_name

        ws['Z213'] = self.seller_director_short_name

        ws['BM213'] = self.payer_director_short_name

        ws['C216'] = self.seller_info

        ws['AS216'] = self.payer_info

        ws['BE7'] = ', '.join(CUR[self.currency.upper()][:-1])

        ws['R10'] = f'ТТН №{self.documents_number} от {self.documents_date}'

        for index, product in enumerate(self.products):
            index += 15
            ws[f'J{index}'] = product[2]
            unit = [i for i in UNITS if product[3] in i][0]
            ws[f'Y{index}'] = unit.split(' ')[0]
            ws[f'W{index}'] = UNITS[unit]
            ws[f'AA{index}'] = '{0:,}'.format(int(product[0])).replace(',', ' ')
            ws[f'AD{index}'] = '{0:,.2f}'.format(float(product[1]), 2).replace(',', ' ').replace('.', ',')
            ws[f'AN{index}'] = '{0:,.2f}'.format(float(product[5]), 2).replace(',', ' ').replace('.', ',')
            ws[f'BF{index}'] = '{0:,.2f}'.format(float(product[5]), 2).replace(',', ' ').replace('.', ',')


        ws['AN192'] = self.total_sum
        ws['BF192'] = self.total_sum

        ws['O201'] = f'ТТН №{self.documents_number} от {self.documents_date}, {self.powerOfAttorney_number} от {self.documents_date}'

        wb.save(self.distance_dir+str(datetime.now().timestamp()).split('.')[0]+'УПД.xlsx')


    def generate_cmr(self):
        cmr_data = {'line_2':f"ТТН №{self.documents_number} от {self.documents_date}, УПД №{self.documents_number} от {self.documents_date}",
            'line_1': f'Контракт №{self.document["contract"]} от {self.document["contract_date"]}, Приложение №{self.document["appendix"]} от {self.document["appendix_date"]}',
            'consignee_country':self.consignee['country'],
            'consignee_street':self.consignee['street'],
            'shipper_country':self.shipper['country'],
            'shipper_name':self.shipper['name'],
            'driver_name':self.powerOfAttorney['driver'],
            'shipper_city':self.shipper['city'],
            'documents_date':self.documents_date,
            'driver_short_name':self.driver_short_name,
            'car':powerOfAttorney['car'],
            'car_number':powerOfAttorney['car_number'],
            'consignee_name':self.consignee['name'],
            'driver_passport':self.powerOfAttorney['passport'],
            'currency_name':CUR[self.currency.upper()][-1],
            'driver_passport_date':self.powerOfAttorney['passport_date'],
            'consignee_city':self.consignee['city'],
            'shipper_street':self.shipper['street'],
            'documents_number':self.documents_number,}

        k = 0
        mal = 0
        cmr_number = randint(1000000, 9999999)
        mal_sum = 0
        for ind in range(len(self.products)):
            cmr_data[f'product_{mal+1}']=f'{self.products[ind][0]} {self.products[ind][4]} {self.products[ind][2]}'
            mal += 1
            mal_sum += float(self.products[ind][5])
            if ind%6==0:
                cmr_data['total']='{0:,.2f}'.format(float(mal_sum), 2).replace(',', ' ').replace('.', ',')
                doc = DocxTemplate(f"{self.shablons_dir}CMR-1 .docx")
                cmr_data['cmr_number']=cmr_number+k
                inner = self.distance_dir+str(datetime.now().timestamp()).split('.')[0]+f'CMR-{k}.docx'
                doc.render(cmr_data)
                doc.save(inner)
                k+=1
                mal = 0
                mal_sum = 0
                cmr_data[f'product_1'] = ''
                cmr_data[f'product_2'] = ''
                cmr_data[f'product_3'] = ''
                cmr_data[f'product_4'] = ''
                cmr_data[f'product_5'] = ''
                cmr_data[f'product_6'] = ''
        cmr_data['total']='{0:,.2f}'.format(float(mal_sum), 2).replace(',', ' ').replace('.', ',')
        doc = DocxTemplate(f"{self.shablons_dir}CMR-1 .docx")
        cmr_data['cmr_number']=cmr_number+k
        inner = self.distance_dir+str(datetime.now().timestamp()).split('.')[0]+f'CMR-{k}.docx'
        doc.render(cmr_data)
        doc.save(inner)

    def generate_dover(self):
        
        data = {
            "company_name":self.seller['name'],
            "company_address":self.seller_addres,
            "company_tin":self.seller['tin'],
            "doc_number":self.powerOfAttorney_number,
            "doc_date":self.documents_date,
            "documents_date":self.documents_date,
            "driver_datv":get_datv_name(powerOfAttorney["driver"]),
            "passport":powerOfAttorney['passport'],
            "passport_date":powerOfAttorney['passport_date'],
            "company_short_name":self.seller['short_name'],
            "documents_number":self.documents_number,
            "consignee":self.consignee["name"],
            "consignee_address":self.consignee_addres
        }
        doc = DocxTemplate(f'{self.shablons_dir}Dover.docx')
        inner = self.distance_dir+str(datetime.now().timestamp()).split('.')[0]+f'Доверенность.docx'
        doc.render(data)
        doc.save(inner)

    def gnerate_transport_request(self):
        products = []
        for product in self.products:
            pr = {
                'product_name_ru':product[2],
                'product_price':'{0:,.2f}'.format(float(product[5]), 2).replace(',', ' ').replace('.', ','),
                'product_price_str':get_sum_str(product[5], self.currency),
            }
            products.append(pr)
        order_date = {'car':powerOfAttorney['car'],
                      'payer_address':self.payer_addres,
                      'driver_full_name':powerOfAttorney['driver'],
                      'date_consignee':'date_consignee',
                      'payer_name':self.payer['name'],
                      'address_shipper':self.shipper_addres,
                      'products':products,
                      'shipper_company':self.shipper['name'],
                      'product_price_total':f'{self.total_sum} {CUR[self.currency.upper()][-1]} ({get_sum_str(self.total_sum, self.currency)})',
                      'country_consignee':self.consignee['country'],
                      'country_shipper':self.shipper['country'],
                      'city_Consignee':self.consignee['city'],
                      'city_shipper':self.shipper['city'],
                      'date_shipper':self.document['date'],
                      'payer_tin':self.payer['tin'],
                      'address_consignee':self.consignee_addres,
                      'car_number':powerOfAttorney['car_number'],
                      'documents_number':self.documents_number,
                      'documents_date':self.documents_date,
                      'consignee_company':self.consignee['name'],
                      }
        doc = DocxTemplate(f"{self.shablons_dir}Транспортная-заявка.docx")
        inner = self.distance_dir+str(datetime.now().timestamp()).split('.')[0]+f'Transport requests.docx'
        doc.render(order_date)
        doc.save(inner)

    def generate_packing_list(self):
        products = []
        x = 1
        for product in self.products:
            pr = {
                'number':x,
                'name':product[2],
                'qtt':'{0:,}'.format(int(product[0])).replace(',', ' '),
                'unit':'{0:,.2f}'.format(float(product[1]), 2).replace(',', ' ').replace('.', ','),
                'total':'{0:,.2f}'.format(float(product[5]), 2).replace(',', ' ').replace('.', ',')
            }
            x+=1
            products.append(pr)
        data = {
            'total':f'{self.total_sum} {CUR[self.currency.upper()][-1]} ({get_sum_str(self.total_sum, self.currency)})', 
            'consignee_name':self.consignee['name'], 
            'contract_date':self.document['contract_date'], 
            'documents_date':self.documents_date, 
            'shipper_address':self.shipper_addres, 
            'contract_number':self.document['contract'], 
            'appendix_number':self.document['appendix'], 
            'shipper_name':self.shipper['name'], 
            'consignee_address':self.consignee_addres, 
            'appendix_date':self.document['appendix_date'], 
            'products':products, 
            'director_short_name':self.shipper_director_short_name, 
            'currency':CUR[self.currency.upper()][0]
            }
        doc = DocxTemplate(f"{self.shablons_dir}packing_list.docx")
        inner = self.distance_dir+str(datetime.now().timestamp()).split('.')[0]+f'Packing list.docx'
        doc.render(data)
        doc.save(inner)


    def generate(self):
        self.generate_ttn()
        self.generate_upd() 
        self.generate_cmr()
        self.gnerate_transport_request()
        self.generate_packing_list()
        self.generate_dover()


if "__main__" == __name__:
    
    doc = Documents()
    doc.generate()
