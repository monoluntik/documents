from openpyxl import load_workbook
from datetime import datetime

wbData = load_workbook(filename = 'dataUsd.xlsx')
wsData = wbData['Table 1']

wbshab = load_workbook(filename = 'Выписка.xlsx')
wsshab = wbshab['Лист1']

wbDataRub = load_workbook(filename = 'dataRub.xlsx')
wsDataRub = wbDataRub['Table 1']

wbDataSom = load_workbook(filename = 'dataSom.xlsx')
wsDataSom = wbDataSom['Table 1']

# print(wsshab)
list_ = ['A', 'B', 'C', 'D', 'F', 'G', 'E', 'H']
list_1 = ['A', 'B', 'C', 'D', 'F', 'G', 'E', 'H', 'I', 'J', 'K', 'L', 'M']


# for x in range(1, 832):
#     wsshab[f'A{x+1}'] = f"{wsData[f'A{x}'].value}".split(' ')[0]
#     wsshab[f'B{x+1}'] = 'РСК'
#     wsshab[f'K{x+1}'] = wsData[f'H{x}'].value

#     if 'Оплата комиссии' in wsData[f'H{x}'].value:
#         wsshab[f'H{x+1}'] = wsData[f'F{x}'].value 

#     elif 'ОАО РСК Банк'.lower() in wsData[f'E{x}'].value.lower():
#         wsshab[f'E{x+1}'] = wsData[f'G{x}'].value

#     else:
#         wsshab[f'F{x+1}'] = wsData[f'F{x}'].value
#         wsshab[f'I{x+1}'] = 'ОсОО Трейд Сервис КейДжи'
        

#     wsshab[f'J{x+1}'] = wsData[f'C{x}'].value
#     wsshab[f'L{x+1}'] = wsData[f'E{x}'].value
#     wsshab[f'G{x+1}'] = "USD"


#     for xl in range(1,205):
#         # print(wsDataRub[f'H{xl}'].value, xl, f'H{xl}')
#         if 'операционные средства' in f"{wsData[f'H{x}'].value}".replace('\n', ' ').lower():
#             wsshab[f'C{x+1}'] = wsData[f'G{x}'].value
#             wsshab[f'D{x+1}'] = 'USD'
#         if f"{wsData[f'H{x}'].value}".replace('\n', ' ') == f"{wsDataRub[f'H{xl}'].value}".replace('\n', ' '):
#             wsshab[f'D{x+1}'] = 'RUB'
#             wsshab[f'C{x+1}'] = wsDataRub[f'F{xl}'].value
#             if wsDataRub[f'F{xl}'].value == '':
#                print('____________') 

#     for xl in range(1,118):
#         # print(wsDataRub[f'H{xl}'].value, xl, f'H{xl}')
#         if 'операционные средства' in f"{wsData[f'H{x}'].value}".replace('\n', ' ').lower():
#             wsshab[f'C{x+1}'] = wsData[f'G{x}'].value
#             wsshab[f'D{x+1}'] = 'USD'
#         if f"{wsData[f'H{x}'].value}".replace('\n', ' ') == f"{wsDataSom[f'H{xl}'].value}".replace('\n', ' '):
#             wsshab[f'D{x+1}'] = 'Som'
#             wsshab[f'C{x+1}'] = wsDataSom[f'F{xl}'].value
#             if wsDataSom[f'F{xl}'].value == '':
#                print('____________') 



BakaiDol = load_workbook(filename = 'интеко долл выписка.xlsx')
BakaiDollar = BakaiDol['W3C Example Table']

BakaiRub = load_workbook(filename = 'Интеко руб.xlsx')
BakaiRubles = BakaiRub['W3C Example Table']

BakaiSom = load_workbook(filename = 'Интеко сом.xlsx')
BakaiKGS = BakaiSom['W3C Example Table']

BakaiEURO = load_workbook(filename = 'Интеко евро выписка.xlsx')
BakaiEURO = BakaiEURO['W3C Example Table']

for i, k in zip(range(2, 357), range(2, 357)):
    wsshab[f'B{i}'] = 'Бакай'
    wsshab[f'K{i}'] = BakaiDollar[f'J{k}'].value
    wsshab[f'A{i}'] = BakaiDollar[f'A{k}'].value
    wsshab[f'G{i}'] = "USD"
    wsshab[f'L{i}'] = BakaiDollar[f'D{k}'].value
    wsshab[f'J{i}'] = BakaiDollar[f'F{k}'].value
    if not 'КЫРГЫЗСТАН' in BakaiDollar[f'E{k}'].value:
        wsshab[f'M{i}'] = BakaiDollar[f'E{k}'].value

    if not 'Комиссия за перевод' in BakaiDollar[f'J{k}'].value and 'Перевод SWIFT' in BakaiDollar[f'J{k}'].value:
        wsshab[f'F{i}'] = BakaiDollar[f'H{k}'].value
        wsshab[f'I{i}'] = 'ОсОО Интеко Транс'
    elif 'Комиссия за перевод' in BakaiDollar[f'J{k}'].value:
        wsshab[f'H{i}'] = BakaiDollar[f'H{k}'].value
    elif 'комиссия' in BakaiDollar[f'J{k}'].value:
        wsshab[f'H{i}'] = BakaiDollar[f'H{k}'].value
    else:
        if 'Конвертация' in BakaiDollar[f'J{k}'].value:
            for pl in range(2, 22):
                if BakaiKGS[f'J{pl}'].value == BakaiDollar[f'J{k}'].value:
                    wsshab[f'C{i}'] = BakaiKGS[f'H{pl}'].value
                    wsshab[f'D{i}'] = 'Som'
                    wsshab[f'E{i}'] = BakaiDollar[f'I{k}'].value

            for pl in range(2, 166):
                if BakaiRubles[f'J{pl}'].value == BakaiDollar[f'J{k}'].value:
                    wsshab[f'C{i}'] = BakaiRubles[f'H{pl}'].value
                    wsshab[f'D{i}'] = 'RUB'
                    wsshab[f'E{i}'] = BakaiDollar[f'I{k}'].value

            for pl in range(2, 76):
                if BakaiEURO[f'J{pl}'].value == BakaiDollar[f'J{k}'].value:
                    wsshab[f'C{i}'] = BakaiEURO[f'H{pl}'].value
                    wsshab[f'D{i}'] = 'EURO'
                    wsshab[f'E{i}'] = BakaiDollar[f'I{k}'].value
        else:
            wsshab[f'C{i}'] = BakaiDollar[f'I{k}'].value
            wsshab[f'D{i}'] = 'USD'

for i, k in zip(range(357, 522), range(2, 166)):
    wsshab[f'A{i}'] = BakaiRubles[f'A{k}'].value
    wsshab[f'K{i}'] = BakaiRubles[f'J{k}'].value
    wsshab[f'L{i}'] = BakaiRubles[f'D{k}'].value
    wsshab[f'M{i}'] = BakaiRubles[f'E{k}'].value
    wsshab[f'G{i}'] = "RUB"
    if not 'Комиссия за перевод' in BakaiRubles[f'J{k}'].value and 'Перевод' in BakaiRubles[f'J{k}'].value:
        wsshab[f'F{i}'] = BakaiRubles[f'H{k}'].value
        wsshab[f'C{i}'] = BakaiRubles[f'I{k}'].value
    elif 'Комиссия за перевод' in BakaiRubles[f'J{k}'].value:
        wsshab[f'H{i}'] = BakaiRubles[f'H{k}'].value
    elif 'комиссия' in BakaiRubles[f'J{k}'].value:
        wsshab[f'H{i}'] = BakaiRubles[f'H{k}'].value
    else:
        if 'Конвертация' in BakaiRubles[f'J{k}'].value:
            wsshab[f'E{i}'] = BakaiRubles[f'H{k}'].value
            pass
        else:
            wsshab[f'C{i}'] = BakaiRubles[f'I{k}'].value
            wsshab[f'D{i}'] = 'RUB'

for i, k in zip(range(521, 596), range(2, 75)):
    wsshab[f'A{i}'] = BakaiEURO[f'A{k}'].value
    wsshab[f'K{i}'] = BakaiEURO[f'J{k}'].value
    wsshab[f'L{i}'] = BakaiEURO[f'D{k}'].value
    wsshab[f'M{i}'] = BakaiEURO[f'E{k}'].value
    wsshab[f'G{i}'] = "EURO"
    if not 'Комиссия за перевод' in BakaiEURO[f'J{k}'].value and 'Перевод' in BakaiEURO[f'J{k}'].value:
        wsshab[f'F{i}'] = BakaiEURO[f'H{k}'].value
        wsshab[f'C{i}'] = BakaiEURO[f'I{k}'].value
    elif 'Комиссия за перевод' in BakaiEURO[f'J{k}'].value:
        wsshab[f'H{i}'] = BakaiEURO[f'H{k}'].value
    elif 'комиссия' in BakaiEURO[f'J{k}'].value:
        wsshab[f'H{i}'] = BakaiEURO[f'H{k}'].value
    else:
        if 'Конвертация' in BakaiEURO[f'J{k}'].value:
            for pl in range(2, 22):
                if BakaiKGS[f'J{pl}'].value == BakaiEURO[f'J{k}'].value:
                    wsshab[f'C{i}'] = BakaiKGS[f'H{pl}'].value
                    wsshab[f'D{i}'] = 'Som'
                    wsshab[f'E{i}'] = BakaiEURO[f'I{k}'].value

            for pl in range(2, 166):
                if BakaiRubles[f'J{pl}'].value == BakaiEURO[f'J{k}'].value:
                    wsshab[f'C{i}'] = BakaiRubles[f'H{pl}'].value
                    wsshab[f'D{i}'] = 'RUB'
                    wsshab[f'E{i}'] = BakaiEURO[f'I{k}'].value

            for pl in range(2, 356):
                if BakaiEURO[f'J{pl}'].value == BakaiEURO[f'J{k}'].value:
                    wsshab[f'C{i}'] = BakaiDollar[f'H{pl}'].value
                    wsshab[f'D{i}'] = 'USD'
                    wsshab[f'E{i}'] = BakaiEURO[f'I{k}'].value
        else:
            wsshab[f'C{i}'] = BakaiEURO[f'I{k}'].value
            wsshab[f'D{i}'] = 'EURO'


wbshab.save(str(datetime.now().timestamp()).split('.')[0]+'Выписка.xlsx')
